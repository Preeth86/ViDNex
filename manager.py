import subprocess
import pickle
import copy
import os
from openpyxl import load_workbook, Workbook
import json
import sys
import time
import requests
import json
import copy
import networkx as nx
import random
import uuid
import tempfile

RYU_BASE_URL = 'http://127.0.0.1:8080'  # Base URL for Ryu's REST API

def execute_substrate_network(args_file, ch):
    with open(args_file, 'r') as file:
        arguments = file.readline().strip().split()
    arguments.append(str(ch))
    arguments.append("SN/SN.topo.pickle")
    print("\nExecuting Substrate Network with arguments:", arguments)
    mininet_script_path = '/home/vnesdn/PycharmProjects/EFraS++/mininet/SN.py'
    env = os.environ.copy()
    env['PYTHONPATH'] = "/home/vnesdn/PycharmProjects/EFraS++/.venv/lib/python3.7/site-packages"
    python_exec = "/home/vnesdn/PycharmProjects/EFraS++/.venv/bin/python"
    command = ['sudo', '-S', sys.executable, mininet_script_path] + arguments
    print("Running command:", " ".join(command))
    subprocess.run(command, env=env)
    print("\nSubstrate Network Execution Completed.\n")

def rollback_embedding(vm_to_server_assignments, path_mappings, servers, graph,old_vnr):
    # Rollback node embedding by releasing CPU resources
    i=1
    map={}
    for cores in old_vnr['vm_cpu_cores']:
        key="VM"+str(i)
        map[key]=cores
        i=i+1

    for vm, server_id in vm_to_server_assignments.items():
        vm_cpu = 0
        if vm in map:
            vm_cpu=map[vm]
        else:
            vm_cpu=random.randint(0,10)

        servers[server_id]['cpu'] += vm_cpu
        servers[server_id]['vms'] = [v for v in servers[server_id]['vms'] if v['vnr_id'] != vm]
        print(f"Rolled back {vm_cpu} CPU units for server {server_id}. New available CPU: {servers[server_id]['cpu']}")

    bw_vals = old_vnr['bandwidth_values']
    for (source_server, target_server, vnr_id), path in path_mappings:
        for ((srv, tgt, vid), bandwidth_path) in path_mappings:
            i=0
            bandwidth_demand=bw_vals[i]
            i=(i+1)%len(bw_vals)

            for i in range(len(path) - 1):
                node1, node2 = path[i], path[i + 1]
                for link in graph['links']:
                    if node1==link['source'] and node2==link['target']:
                        link['bandwidth']+=bandwidth_demand

                print(f"Rolled back {bandwidth_demand} bandwidth on link {node1} <-> {node2}. ")
    print("Rollback of node and link embedding completed.")

def load_sn_graph(sn_json):
    try:
        graph = nx.Graph()
        servers = {}

        # Add nodes to graph
        for node_id, node_data in sn_json.items():
            if node_id.startswith('h'):
                servers[node_id] = {
                    'cpu': node_data.get('allocated_cores', 0),
                    'original_cpu': node_data.get('original_cores', 0),
                    'vms': []
                }
                graph.add_node(node_id, type='host', cpu=node_data.get('allocated_cores', 0))
            elif node_id.startswith('l') or node_id.startswith('s'):
                if isinstance(node_data, dict):
                    graph.add_node(node_id, type='switch', **node_data)
                else:
                    graph.add_node(node_id, type='switch')

        # Add edges (links between switches and servers)
        for link in sn_json.get('links_details', []):
            node1, node2 = link.get('node1'), link.get('node2')
            bandwidth = link.get('assigned_bandwidth', 0)
            graph.add_edge(node1, node2, bandwidth=bandwidth)

        # Ensure the controller connects to all spine switches
        graph.add_node('c0', type='controller')
        for node in graph.nodes:
            if node.startswith('s'):  # Connect to all spine switches
                graph.add_edge('c0', node, bandwidth=10000)  # High bandwidth for controller connections

        # Ensure all nodes from SN JSON are present
        for node_id in sn_json.keys():
            if node_id not in graph.nodes:
                graph.add_node(node_id)

        link_flags = {edge: False for edge in graph.edges()}
        return servers

    except KeyError as e:
        raise ValueError(f"Invalid SN JSON format. Missing key: {e}")
    except TypeError as e:
        raise ValueError(f"Type error while processing SN JSON: {e}")

def execute_vnr_generator(args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, sz, vl_prob):
    with open(args_file, 'r') as file:
        arguments = file.readline().strip().split()
    python_exec = "/home/vnesdn/PycharmProjects/EFraS++/.venv/bin/python"
    print("Executing VNR with arguments", arguments)
    subprocess.run([python_exec, "mininet/VNR.py", "SN/SN.topo.pickle", vnr_file, sz] + arguments + [vnr_gen_ch, str(vl_prob)])
    print("VNR execution completed.")

def send_sn_to_ryu(sn_pickle_file):
    """Send Substrate Network pickle file to ryu.py, ensuring required attributes are added."""
    with open(sn_pickle_file, 'rb') as f:
        sn_data = pickle.load(f)

    # Ensure each server node has the required 'cpu' and 'original_cpu' attributes
    for node in sn_data.get('nodes', []):
        if node.get('type') == 'server':  # Check if the node is a server
            if 'cpu' not in node:
                node['cpu'] = 1000  # Default CPU value (adjust as needed)
            if 'original_cpu' not in node:
                node['original_cpu'] = node['cpu']  # Set original_cpu to match the initial cpu

    response = requests.post(f"{RYU_BASE_URL}/topology", json=sn_data)
    if response.status_code != 200:
        raise RuntimeError(f"Failed to send SN topology to Ryu. Status code: {response.status_code}, Response: {response.text}")

def fetch_sn_from_ryu():
    """Fetch the Substrate Network graph from ryu.py, including bandwidth, latency, and packet loss."""
    response = requests.get(f"{RYU_BASE_URL}/topology")
    if response.status_code == 200:
        topology_data = response.json()

        # Fetch RTT and packet loss statistics separately
        rtt_response = requests.get(f"{RYU_BASE_URL}/rtt")
        packet_loss_response = requests.get(f"{RYU_BASE_URL}/packet_loss")

        if rtt_response.status_code == 200 and packet_loss_response.status_code == 200:
            topology_data['rtt'] = rtt_response.json()
            topology_data['packet_loss'] = packet_loss_response.json()

        return topology_data
    else:
        raise RuntimeError(f"Failed to fetch SN graph from Ryu. Status code: {response.status_code}")

def send_vnr_to_ryu(vnr_pickle_file):
    """Send VNR pickle file to ryu.py."""
    with open(vnr_pickle_file, 'rb') as f:
        vnr_data = pickle.load(f)
    payload = {"vnr_file": vnr_pickle_file, "vnr_data": vnr_data}
    response = requests.post(f"{RYU_BASE_URL}/vnr", json=payload)
    if response.status_code != 200:
        raise RuntimeError(f"Failed to send VNR data to Ryu. Status code: {response.status_code}, Response: {response.text}")

def fetch_vnrs_from_ryu():
    """Fetch VNR network graphs from ryu.py."""
    response = requests.get(f"{RYU_BASE_URL}/vnr")
    if response.status_code == 200:
        vnrs_data = response.json()  # List of VNR graphs
        return [vnrs_data[i] for i in range(len(vnrs_data))]  # Ensure individual graphs are processed
    else:
        raise RuntimeError(f"Failed to fetch VNR graphs from Ryu. Status code: {response.status_code}, Response: {response.text}")

def load_network_data(path):
    with open(path, 'rb') as f:
        data = pickle.load(f)
    return data

def calculate_total_bandwidth(topology_data):
    try:
        link_bandwidths = {}  # Dictionary to store bandwidth per link

        if 'links_details' in topology_data:
            total_bandwidth = sum(link.get('assigned_bandwidth', 0) for link in topology_data['links_details'])
            for link in topology_data['links_details']:
                node1, node2 = link.get('node1'), link.get('node2')
                bandwidth = link.get('assigned_bandwidth', 0)
                link_bandwidths[(str(node1), str(node2))] = bandwidth  # Convert tuple keys to string
            return total_bandwidth, link_bandwidths

        elif 'links' in topology_data and 'nodes' in topology_data:
            G = nx.node_link_graph(topology_data)
            total_bandwidth = sum(G[u][v].get('bandwidth', 0) for u, v in G.edges)
            for u, v in G.edges:
                link_bandwidths[(str(u), str(v))] = G[u][v].get('bandwidth', 0)
            return total_bandwidth, link_bandwidths

        elif isinstance(topology_data, dict):
            total_bandwidth = sum(
                data.get('bandwidth', 0)
                for u in topology_data
                for v, data in topology_data[u].items()
                if isinstance(data, dict)
            )
            for u in topology_data:
                for v, data in topology_data[u].items():
                    if isinstance(data, dict) and 'bandwidth' in data:
                        link_bandwidths[(str(u), str(v))] = data['bandwidth']
            return total_bandwidth, link_bandwidths

        else:
            raise KeyError("No valid link data found in topology_data.")

    except Exception as e:
        print(f"Error in calculating bandwidth: {e}")
        return 0, {}

def process_topology_data(topology_data):
    link_flags = {}
    node_flags = {}

    # Check and process links
    if 'links_details' in topology_data:
        for link in topology_data['links_details']:
            node1, node2 = link['node1'], link['node2']
            link_flags[(node1, node2)] = False
    else:
        print("Warning: 'links_details' key not found in topology_data.")

    # Check and process nodes
    if topology_data:
        for node_name in topology_data:
            if isinstance(node_name, str) and node_name.startswith('h'):
                node_flags[node_name] = False
    else:
        print("Warning: 'nodes' key not found in topology_data.")

    return link_flags, node_flags

def print_vnr_details(vnr_data, vnr_id=None):
    for vnr in vnr_data:
        if vnr_id is not None and vnr.get('vnr_id') != vnr_id:
            continue

        # Print VNR ID
        print(f"VNR{vnr['vnr_id'] + 1} Details:")

        # Print VM details
        print("VM's details:")
        for i, cpu_cores in enumerate(vnr.get('vm_cpu_cores', []), start=1):
            print(f"VM{i} - CPU Demand: {cpu_cores}")

        # Print Virtual Links details
        print("\nVirtual Links details:")
        for idx, (vm_link, bandwidth) in enumerate(zip(vnr.get('vm_links', []), vnr.get('bandwidth_values', []))):
            if isinstance(vm_link, list) and len(vm_link) == 2:  # Ensure proper link format
                vm1, vm2 = vm_link
                print(f"VM{vm1 + 1} <--> VM{vm2 + 1}, Bandwidth demand: {bandwidth}")
            else:
                print(f"Warning: Invalid link format at index {idx}. Link: {vm_link}")

def extract_vm_to_host(data):
    print("Data received in extract_vm_to_host:", data)  # Debugging line
    if isinstance(data, list):
        vm_to_host = data[0]
        return {vm: host for vm, host in vm_to_host}
    else:
        raise ValueError("Expected data to be a list, but got something else.")

def extract_connections(data):
    connections = data[1]
    return [(conn[0], conn[1], conn[2]) if len(conn) == 3 else (conn[0], conn[1], 0) for conn in connections]


def append_data_to_excel(excel_file_path, data, name):
    if os.path.exists(excel_file_path):
        book = load_workbook(excel_file_path)
    else:
        book = Workbook()

    sheet = book.active
    if sheet.max_row == 1:
        headers = ["S.No", "Algorithm"] + list(data.keys())
        sheet.append(headers)

    row = [sheet.max_row + 1, name]

    for key, value in data.items():
        if isinstance(value, dict):
            row.append(json.dumps(value))  # Convert dictionary values to JSON string
        else:
            row.append(value)

    sheet.append(row)
    book.save(excel_file_path)

def reconstruct_sn_graph(nx_graph):
    sn_reconstructed = {}

    # Rebuild node information
    for node, data in nx_graph.nodes(data=True):
        sn_reconstructed[node] = {key: value for key, value in data.items()}

    # Rebuild link information
    sn_reconstructed['links_details'] = []
    for node1, node2, data in nx_graph.edges(data=True):
        link_info = {
            "node1": node1,
            "node2": node2,
            "assigned_bandwidth": data.get('bandwidth', 0)
        }
        sn_reconstructed['links_details'].append(link_info)

    return sn_reconstructed

vnr_graph_counter = 1

def algo(sn_graph, algo, vnr_graphs, excel_file_path, name, vl_prob):
    global vnr_graph_counter  # Use the global counter
    start_time = time.time()  # Start the timer

    link_flags, node_flags = process_topology_data(sn_graph)
    vnr_count = 0
    vm_count = 0
    Revenue = []
    cost = [sum(info.get('allocated_cores', 0) for node_name, info in sn_graph.items() if node_name.startswith('h'))]

    # Initial State Metrics
    initial_total_bw = calculate_total_bandwidth(sn_graph)
    initial_total_cpu = sum(info['allocated_cores'] for node_name, info in sn_graph.items() if node_name.startswith('h'))
    initial_rtt = sn_graph.get('rtt', {})
    initial_packet_loss = sn_graph.get('packet_loss', {})

    used_servers = set()
    vm_to_server_mapping = {}
    idle_servers = {node_name for node_name in sn_graph if node_name.startswith('h')}
    for vnr_graph in vnr_graphs:
        print(f"\nProcessing VNR Graph {vnr_graph_counter}...")
        # Reset the counters for each VNR Graph to avoid cumulative aggregation
        s_vnr_count = 0
        total_vms_used_for_embedding = 0
        total_vls_used_for_embedding = 0
        total_virtual_links = 0
        total_path_length = 0
        total_paths = 0

        vnr_info = {f'VNR{vnr["vnr_id"]}': {f'VM{i}': {'cpu': cpu} for i, cpu in enumerate(vnr['vm_cpu_cores'], start=1)}
            for vnr in vnr_graph}

        vm_count += sum(len(vnr['vm_cpu_cores']) for vnr in vnr_graph)
        total_virtual_links += sum(len(vnr['vm_links']) for vnr in vnr_graph)
        Revenue.extend([0] * len(vnr_graph))
        cost.extend([0] * len(vnr_graph))
        vnr_count += len(vnr_graph)

        # Create temporary JSON files to store SN and VNR data
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json') as sn_file:
            json.dump(sn_graph, sn_file)
            sn_filepath = sn_file.name

        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json') as vnr_file:
            json.dump(vnr_graph, vnr_file)
            vnr_filepath = vnr_file.name

        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json') as vnr_info_file:
            json.dump(vnr_info, vnr_info_file)
            vnr_info_filepath = vnr_info_file.name

        args = ["python3", algo, vnr_info_filepath, sn_filepath, str(vnr_graph_counter), vnr_filepath]
        vnr_graph_counter += 1
        try:
            subprocess.run(args, check=True)
        except subprocess.CalledProcessError as e:
            print(f"Error executing {algo}: {e}")
        finally:
            # **Clean up temp files after execution**
            os.remove(sn_filepath)
            os.remove(vnr_filepath)
            os.remove(vnr_info_filepath)

        try:
            with open('Node_Link_Embedding_Details.json', 'r') as f:
                embedding_results = json.load(f)

            initial_total_bw, _ = calculate_total_bandwidth(sn_graph)
            initial_total_cpu = sum(info['allocated_cores'] for node_name, info in sn_graph.items() if node_name.startswith('h'))

            for vnr in vnr_graph:
                t = next((item for graph in embedding_results for item in graph['vnr_results'] if item['vnr_id'] == vnr['vnr_id']), None)
                if not t or 'embedding_success' not in t or not t['embedding_success'].get(str(vnr['vnr_id']), False):
                    continue

                s_vnr_count += 1
                total_vms_used_for_embedding += len(t['vm_to_server_assignments'])
                total_vls_used_for_embedding += len(t['path_mappings'])

                for path in t['path_mappings']:
                    total_path_length += len(path[1])
                total_paths += len(t['path_mappings'])

                if 'updated_graph' in t:
                    updated_graph = nx.node_link_graph(t['updated_graph'])
                    sn_graph = reconstruct_sn_graph(updated_graph)

                for link in t['link_flags']:
                    link_flags[tuple(link[0])] = link[1]

                for vm, server in t['vm_to_server_assignments'].items():
                    node_flags[server] = True
                    used_servers.add(server)
                    vm_to_server_mapping[vm] = server

                reve = sum(vnr['vm_cpu_cores']) + sum(vnr['bandwidth_values'])
                cos = sum(vnr['vm_cpu_cores'])
                Revenue[vnr['vnr_id'] - 1] = reve
                cost[vnr['vnr_id'] - 1] = cos

            # Performance matrix calculation for each VNR graph
            final_total_bw, final_bandwidth_per_link = calculate_total_bandwidth(sn_graph)
            final_bandwidth_per_link_str = json.dumps({str(k): v for k, v in final_bandwidth_per_link.items()})
            after_embedding_total_cpu = sum(info['allocated_cores'] for node_name, info in sn_graph.items() if node_name.startswith('h'))
            updated_sn_graph = fetch_sn_from_ryu()
            final_total_cpu = sum(info['allocated_cores'] for node_name, info in updated_sn_graph.items() if node_name.startswith('h'))
            final_rtt = updated_sn_graph.get('rtt', {})
            final_packet_loss = updated_sn_graph.get('packet_loss', {})

            # Performance Metrics Calculation
            avg_rtt_before = sum(initial_rtt.values()) / len(initial_rtt) if initial_rtt else 0
            avg_rtt_after = sum(final_rtt.values()) / len(final_rtt) if final_rtt else 0
            avg_packet_loss_before = sum(initial_packet_loss.values()) / len(initial_packet_loss) if initial_packet_loss else 0
            avg_packet_loss_after = sum(final_packet_loss.values()) / len(final_packet_loss) if final_packet_loss else 0

            AR = round((s_vnr_count / vnr_count) * 100, 2) if vnr_count > 0 else 0
            TS = len(node_flags)
            SU = sum(1 for flag in node_flags.values() if flag is True)
            PL = len(link_flags)
            LU = sum(1 for flag in link_flags.values() if flag is True)
            NS = round((total_vms_used_for_embedding / TS), 2) if TS > 0 else 0
            ANS = round((total_vms_used_for_embedding / SU), 2) if SU > 0 else 0
            LS = round((total_vls_used_for_embedding / PL), 2) if PL > 0 else 0
            ALS = round((total_vls_used_for_embedding / LU), 2) if LU > 0 else 0

            print('\n\033[1m\033[4m' + "Performance Matrices calculations" + '\033[0m')
            print(f"Acceptance Ratio: {AR}% (Out of {vnr_count} VNRs {s_vnr_count} VNRs are accepted)")
            print(f"Total Servers: {TS}, Servers Used: {SU}, Idle Servers: {TS - SU}")
            print(f"Total Physical Links: {PL}, Links Used: {LU}, Idle Links: {PL - LU}")
            print(f"Total Virtual Links: {total_virtual_links}")
            print(f"Node Stress: {NS}, Link Stress: {LS}")
            print(f"Active Nodes Stress: {ANS}, Active Link Stress: {ALS}")

            # Calculate energy consumption
            O_SN_host = {}
            energy_consu = {}
            P_idle = 150
            P_full = 300

            for host, info in sn_graph.items():
                if host.startswith('h'):
                    O_SN_host[host] = info['original_cores']

            UP_SN_host = {}
            for host, info in sn_graph.items():
                if host.startswith('h'):
                    UP_SN_host[host] = info['allocated_cores']
                    energy_consu[host] = round(P_idle + (P_full - P_idle) * ((O_SN_host[host] - UP_SN_host[host]) / O_SN_host[host]), 2)

            total_energy_used_sn = round(sum(energy_consu[host] for host, used in node_flags.items() if used), 2)
            average_path_length = round(total_path_length / total_paths, 2) if total_paths > 0 else 0

            # **Load Balancing Efficiency Calculation**
            cpu_usages = [server['cpu'] for server in sn_graph.values() if isinstance(server, dict) and server.get('type') == 'host']
            max_load = max(cpu_usages)
            min_load = min(cpu_usages)
            epsilon = 1e-6  # Small constant to prevent division by zero
            load_balance_efficiency = 1 - (max_load - min_load) / (max_load + epsilon)

            data = {
                "VNR ID": ", ".join([f"VNR{vnr['vnr_id']}" for graph in vnr_graphs for vnr in graph]),
                "VL Connectivity": vl_prob,
                "Total Number of VNRs": vnr_count,
                "Accepted Number of VNRs": s_vnr_count,
                "Acceptance Ratio": f"{AR}%",
                "Total Available Servers": TS,
                "Number of Servers Used": SU,
                "Names of Servers Used": ", ".join(sorted(used_servers)),
                "Number of Idle Servers": TS - SU,
                "Names of Idle Servers": ", ".join(sorted(idle_servers)),
                "VM-to-Server Mapping": ", ".join([f"{vm} -> {server}" for vm, server in vm_to_server_mapping.items()]),
                "Total Available Physical Links": PL,
                "Number of Links Used": LU,
                "Number of Idle Links": PL - LU,
                "Total Number of VMs": vm_count,
                "Total VMs embedded": total_vms_used_for_embedding,
                "Total Number of VLs": total_virtual_links,
                "Total VLs embedded": total_vls_used_for_embedding,
                "Node Stress":NS,
                "Active Node Stress":ANS,
                "Link Stress":LS,
                "Active Link Stress":ALS,
                "Before Embedding Total Available CPU of SN": initial_total_cpu,
                "After Embedding Total Available CPU of SN": after_embedding_total_cpu,
                "Before Embedding Total Available BW of SN": initial_total_bw,
                "After Embedding Total Available BW of SN": final_total_bw,
                "Average Path Length": average_path_length,
                "Before Embedding RTT": avg_rtt_before,
                "After Embedding RTT": avg_rtt_after,
                "Before Embedding Packet Loss": avg_packet_loss_before,
                "After Embedding Packet Loss": avg_packet_loss_after,
                "Link Bandwidth Details": final_bandwidth_per_link_str,
                "Total Energy of Embedded servers only in SN": f"{total_energy_used_sn} Watts",
                "Load Balance Efficiency": load_balance_efficiency
            }

            total_ratio = 0
            for i in range(len(Revenue)):
                if Revenue[i] != 0 and cost[i] != 0:
                    total_ratio += Revenue[i] / cost[i]
            avg_rc_ratio = round(total_ratio / s_vnr_count, 2) if s_vnr_count > 0 else "N/A"
            data["Avg R/C Ratio"] = avg_rc_ratio

            Total_energy = round(sum(energy_consu[host] for host in energy_consu), 2)
            data["Total Energy of SN"] = f"{Total_energy} Watts"

            total_execution_time = round(time.time() - start_time, 2)
            data["Total Execution Time"] = f"{total_execution_time} seconds"

            append_data_to_excel(excel_file_path, data, name)

        except (FileNotFoundError, json.JSONDecodeError):
            print("Embedding result file not found or invalid.")
            continue

def main():
    global new_cores
    excel_file_path = 'OUTPUT/parameters_output.xlsx'

    print("Enter your choice for SN and VM Distribution:")
    print("1. Random Distribution")
    print("2. Uniform Distribution")
    print("3. Normal Distribution")
    print("4. Poisson Distribution")
    sn_vm_gen_ch = input().strip()

    args_file = 'SN-Input-File.txt'
    execute_substrate_network(args_file, sn_vm_gen_ch)
    substrate_pickle_file_path = 'SN/SN.topo.pickle'

    send_sn_to_ryu(substrate_pickle_file_path)
    sn_graph = fetch_sn_from_ryu()
    print(sn_graph)
    if not sn_graph:
        print("Failed to fetch SN graph from Ryu. Exiting.")
        return

    VNRs = []

    while True:
        print("Enter your choice for VNR Distribution:")
        print("1. Random Distribution")
        print("2. Uniform Distribution")
        print("3. Normal Distribution")
        print("4. Poisson Distribution")
        vnr_gen_ch = input().strip()

        print("Select Virtual Link Connectivity Probability:")
        print("1. 0.4")
        print("2. 0.9")
        vl_prob_choice = input().strip()

        vl_prob = 0.4 if vl_prob_choice == '1' else 0.9

        print("Enter your choice for Number of VNRs:")
        print("1. 200 VNRs")
        print("2. 400 VNRs")
        print("3. 600 VNRs")
        print("4. 800 VNRs")
        print("5. 1000 VNRs")
        print("0. Exit")
        try:
            vnr_ch = int(input().strip())
        except ValueError:
            print("Invalid input. Please enter a number between 0 and 5.")
            continue

        if vnr_ch == 0:
            break

        vnr_args_file = "VNE-Input-File.txt"
        vnr_file = ""
        for i in range(10):
            if vnr_ch == 1:
                vnr_file = f"VNR/vnr200.{i}.topo.pickle"
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '200', str(vl_prob))
                VNRs.append(vnr_file)
            elif vnr_ch == 2:
                vnr_file = f"VNR/vnr400.{i}.topo.pickle"
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '400', str(vl_prob))
                VNRs.append(vnr_file)
            elif vnr_ch == 3:
                vnr_file = f"VNR/vnr600.{i}.topo.pickle"
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '600', str(vl_prob))
                VNRs.append(vnr_file)
            elif vnr_ch == 4:
                vnr_file = f"VNR/vnr800.{i}.topo.pickle"
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '800', str(vl_prob))
                VNRs.append(vnr_file)
            elif vnr_ch == 5:
                vnr_file = f"VNR/vnr1000.{i}.topo.pickle"
                execute_vnr_generator(vnr_args_file, vnr_gen_ch, sn_vm_gen_ch, vnr_file, '1000', str(vl_prob))
                VNRs.append(vnr_file)
            else:
                print("Invalid choice for Number of VNRs. Please select between 1 and 5.")
                break

        for vnr in VNRs:
            send_vnr_to_ryu(vnr)

        vnr_graphs = fetch_vnrs_from_ryu()

        if not vnr_graphs:
            print("Failed to fetch VNR graphs from Ryu. Exiting.")
            return

        while True:
            print("Select the Algorithm to execute:")
            print("1. NEXA")
            print("2. DQL")
            print("3. LitE")
            print("4. CEVNE")
            print("5. DROI")
            print("6. First Fit")
            print("7. SCA-R")
            print("8. NORD")
            print("0. Go back to Number of VNRs")
            try:
                algo_ch = int(input().strip())
            except ValueError:
                print("Invalid input. Please enter 1 to 8.")
                continue

            if algo_ch == 0:
                break
            elif algo_ch == 1:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "Lite++.py", [vnr_graph], excel_file_path, "NEXA", str(vl_prob))
                print("NEXA Algorithm Execution Completed.")

                while True:
                    print("What scaling options would you like to perform to the servers:")
                    print("1.Scale up")
                    print("2.Scale Down")
                    print("0.None of the above")

                    scaler = int(input().strip())
                    if scaler==0:
                        break
                    elif scaler==1:
                        random_nos_map={}
                        st=1
                        vnr_cnt=0
                        n=len(vnr_graphs)
                        print(vnr_graphs)
                        for i in range(5):
                            random_nos_map[random.randint(0,n)]=i
                        print(random_nos_map)

                        for vnr_graph in vnr_graphs:
                                for vnr in vnr_graph:
                                        old_vnr=copy.deepcopy(vnr)
                                        new_cpu_cores=[]
                                        id = vnr['vnr_id']
                                        if id==1:
                                            for val in vnr['vm_cpu_cores']:
                                                new_cpu_cores.append(val*1.2)
                                            vnr['vm_cpu_cores']=new_cpu_cores
                                        pid = vnr['vnr_id']
                                        with open('Node_Link_Embedding_Details.json', 'r') as f:
                                            embedding_results = json.load(f)
                                        vnr_result = next(
                                            item for item in embedding_results[0]['vnr_results'] if item['vnr_id'] == id)

                                        path_mappings = vnr_result['path_mappings']
                                        vm_to_server_assignments = vnr_result['vm_to_server_assignments']
                                        updated_graph = vnr_result['updated_graph']
                                        print("Path Mappings:", path_mappings)
                                        print("vnr:",vnr)
                                        print(vm_to_server_assignments)

                                        servers=load_sn_graph(sn_graph)

                                        rollback_embedding(vm_to_server_assignments,path_mappings,servers,updated_graph,old_vnr)
                                        algo(sn_graph, "Lite++.py", [vnr_graph], excel_file_path, "Lite++")
                                        vnr_cnt+=1
                                        if vnr_cnt==n:
                                            break
                                if vnr_cnt==n:
                                    break

                        print("Lite++ Algorithm Upscaled")

                    elif scaler==2:
                        random_nos_map={}
                        n= len(vnr_graphs)
                        st=1
                        vnr_cnt=0
                        print(vnr_graphs)
                        for i in range(5):
                            random_nos_map[random.randint(0,n)]=i
                        print(random_nos_map)

                        for vnr_graph in vnr_graphs:
                                for vnr in vnr_graph:
                                        old_vnr=copy.deepcopy(vnr)
                                        new_cpu_cores=[]
                                        id = vnr['vnr_id']
                                        if id == 1:
                                            for val in vnr['vm_cpu_cores']:
                                                new_cpu_cores.append(val * 0.8)
                                            vnr['vm_cpu_cores'] = new_cpu_cores
                                        vnr['vm_cpu_cores']=new_cpu_cores
                                        pid = vnr['vnr_id']
                                        with open('Node_Link_Embedding_Details.json', 'r') as f:
                                            embedding_results = json.load(f)
                                        vnr_result = next(
                                            item for item in embedding_results[0]['vnr_results'] if item['vnr_id'] == id)

                                        path_mappings = vnr_result['path_mappings']
                                        vm_to_server_assignments = vnr_result['vm_to_server_assignments']
                                        updated_graph = vnr_result['updated_graph']
                                        print("Path Mappings:", path_mappings)
                                        print("vnr:", vnr)
                                        print(vm_to_server_assignments)

                                        servers=load_sn_graph(sn_graph)

                                        rollback_embedding(vm_to_server_assignments,path_mappings,servers,updated_graph,old_vnr)
                                        algo(sn_graph, "Lite++.py", [vnr_graph], excel_file_path, "Lite++")
                                        vnr_cnt+=1
                                        if vnr_cnt==n:
                                            break
                                if vnr_cnt==n:
                                    break

                        print("Lite++ Algorithm Downscaled")

            elif algo_ch==2:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/DQL/test_vne.py", [vnr_graph], excel_file_path, "DQL", str(vl_prob))
                print("DQL Algorithm Execution Completed.")
            elif algo_ch == 3:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/Lite.py", [vnr_graph], excel_file_path, "Lite", str(vl_prob))
                print("Lite Algorithm Execution Completed.")
            elif algo_ch == 4:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/CEVNE.py", [vnr_graph], excel_file_path, "CEVNE", str(vl_prob))
                print("CEVNE Algorithm Execution Completed.")
            elif algo_ch == 5:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/DROI.py", [vnr_graph], excel_file_path, "DROI", str(vl_prob))
                print("DROI Algorithm Execution Completed.")
            elif algo_ch == 6:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/First_Fit.py", [vnr_graph], excel_file_path, "First Fit", str(vl_prob))
                print("First Fit Algorithm Execution Completed.")
            elif algo_ch == 7:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/SCA-R.py", [vnr_graph], excel_file_path, "SCA-R", str(vl_prob))
                print("SCA-R Algorithm Execution Completed.")
            elif algo_ch == 8:
                for vnr_graph in vnr_graphs:
                    algo(sn_graph, "/home/vnesdn/PycharmProjects/EFraS++/NORD.py", [vnr_graph], excel_file_path, "NORD", str(vl_prob))
                print("NORD Algorithm Execution Completed.")
            else:
                print("Invalid choice, please select a valid algorithm.")

if __name__ == "__main__":
    main()
